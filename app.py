import streamlit as st
import openpyxl
from google import genai
from google.genai import types
import json
import io

GOOGLE_API_KEY = "AQ.Ab8RN6IniXQrGk1AxwECnzfKIpFKWqtSx3F3_A58nCXqhUow4g"
client = genai.Client(api_key=GOOGLE_API_KEY)

st.set_page_config(page_title="ยังคิดชื่อเฟี้ยวๆไม่ออก", page_icon="📊", layout="centered")
                   
st.title("ยังคิดชื่อเฟี้ยวๆไม่ออก")
st.write("อัปโหลดไฟล์งบการเงิน PDF เพื่อนำข้อมูลไปอัปเดตลงในไฟล์ Excel อัตโนมัติ")

excel_template = st.file_uploader("อัปโหลดไฟล์ Code_CS_full_financials", type=["xlsx"])
uploaded_files = st.file_uploader("อัปโหลดไฟล์ PDF (Ratio และ Cash Flow)", type=["pdf"], accept_multiple_files=True)


PROMPT_5_YEARS_THAI = """
คุณคือผู้เชี่ยวชาญด้านการบัญชีและวิเคราะห์งบการเงิน หน้าที่ของคุณคืออ่านไฟล์ PDF ข้อมูลทางการเงินที่แนบมานี้ 
แล้วดึงตัวเลขของหัวข้อที่กำหนด ย้อนหลังทั้งหมด 5 ปี โดยเรียงลำดับจากปีล่าสุดไปสู่อดีต ได้แก่ ปี 2568, 2567, 2566, 2565 และ 2564 อย่างเคร่งครัด

หัวข้อทางการเงินที่ต้องการให้ดึงข้อมูล (ดึงเฉพาะตัวเลข):
1. อัตรากำไรสุทธิ(%)
2. อัตราส่วนทุนหมุนเวียน(เท่า)
3. อัตราส่วนทุนหมุนเวียนเร็ว(เท่า)
4. เงินสดสุทธิได้มาจาก(ใช้ไป)กิจกรรมดำเนินงาน
5. อัตราส่วนหนี้สินต่อส่วนของผู้ถือหุ้น(เท่า)
6. ระยะเก็บหนี้(วัน)
7. ระยะเวลาขายสินค้าเฉลี่ย(วัน)
8. ระยะเวลาชำระหนี้(วัน)
9. อัตราการเพิ่มของรายได้รวม(%)
10. อัตราการเพิ่มของกำไรสุทธิ(%)
11. อัตราผลตอบแทนจากสินทรัพย์(%)

ข้อบังคับในการตรวจสอบและคัดแยกตัวเลข (สำคัญมาก):
- เข้มงวดเรื่องค่าว่าง: หากในตารางงบการเงินของปีใดปรากฏเครื่องหมายแดช "-", หรือปล่อยว่างไว้ ซึ่งหมายถึงไม่มีตัวเลขในปีนั้น ให้ใส่ค่านั้นเป็น null ทันที **ห้ามไปหยิบเอาตัวเลขของปีข้างๆ หรือของอุตสาหกรรม หรือบรรทัดอื่นมาใส่แทนเด็ดขาด** ข้อมูลของปีไหนต้องตรงกับคอลัมน์ของปีนั้นใน PDF เท่านั้น
- เข้มงวดเรื่องค่าติดลบ: ให้ตรวจสอบตัวเลขให้ดี หากตัวเลขใดมีเครื่องหมายด้านหน้า "-" ให้แปลงเป็นค่าติดลบในรูปแบบตัวเลขทางคอมพิวเตอร์เสมอ (เช่น -53.24 หรือ -12.43) **ห้ามเอาเครื่องหมายลบออกเด็ดขาด**
- ให้ตรวจสอบข้อมูลแยกตามรายปี (2568, 2567, 2566, 2565, 2564)
- หากปีใด หรือหัวข้อใด ไม่มีข้อมูลปรากฏในไฟล์ PDF ให้ใส่ค่านั้นเป็น null ห้ามเดาตัวเลขเด็ดขาด
- ให้ตอบกลับในรูปแบบ JSON Object เท่านั้น ห้ามมีคำอธิบายอื่น 
- คีย์ (Key) ของหัวข้อต้องสะกดตรงตามภาษาไทย 11 ข้อด้านบนเป๊ะๆ และคีย์ของปีต้องเป็น "2568", "2567", "2566", "2565", "2564"
- ตัวเลขที่ส่งกลับมาต้องห้ามมีเครื่องหมายคอมม่า (,) หรือเครื่องหมายเปอร์เซ็นต์ (%) ให้ส่งเป็นตัวเลขล้วนเท่านั้น
"""

if st.button("Start") and uploaded_files and excel_template:
    with st.spinner("Preparing with AI..."):
        
        pdf_parts = []
        for uploaded_file in uploaded_files:
            file_bytes = uploaded_file.read()
            pdf_parts.append(
                types.Part(
                    inline_data=types.Blob(
                        data=file_bytes,
                        mime_type="application/pdf")))

        try:
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=pdf_parts + [PROMPT_5_YEARS_THAI],
                config=types.GenerateContentConfig(
                    response_mime_type="application/json"
                )
            )
            ai_extracted_json_thai = json.loads(response.text)
            st.success("Done! AI ประมวลผลสกัดข้อมูลสำเร็จ")
        except Exception as e:
            st.error("Error, Try Again")
            st.caption(f"{str(e)}")
            st.stop()

    with st.spinner("กำลังหยอดตัวเลขลงช่องใน Excel แยกตามปี..."):
        wb = openpyxl.load_workbook(excel_template)
        ws = wb.active 
        
    
        column_map_year = {
            "2568": "D",
            "2567": "F",
            "2566": "H",
            "2565": "J",
            "2564": "L"
        }
        
    
        row_map_metric = {
            "อัตรากำไรสุทธิ(%)": 3,
            "อัตราส่วนทุนหมุนเวียน(เท่า)": 4,
            "อัตราส่วนทุนหมุนเวียนเร็ว(เท่า)": 5,
            "เงินสดสุทธิได้มาจาก(ใช้ไป)กิจกรรมดำเนินงาน": 6,
            "อัตราส่วนหนี้สินต่อส่วนของผู้ถือหุ้น(เท่า)": 7,
            "ระยะเก็บหนี้(วัน)": 8,
            "ระยะเวลาขายสินค้าเฉลี่ย(วัน)": 9,
            "ระยะเวลาชำระหนี้(วัน)": 10,
            "อัตราการเพิ่มของรายได้รวม(%)": 11,
            "อัตราการเพิ่มของกำไรสุทธิ(%)": 12,
            "อัตราผลตอบแทนจากสินทรัพย์(%)": 13
        }
        
        
        for first_key, second_data in ai_extracted_json_thai.items():
            if first_key in column_map_year:
                col_letter = column_map_year[first_key]
                if isinstance(second_data, dict):
                    for metric_name, val in second_data.items():
                        if metric_name in row_map_metric:
                            row_num = row_map_metric[metric_name]
                            cell_target = f"{col_letter}{row_num}"
                            ws[cell_target] = "" if val is None else val

            
            elif first_key in row_map_metric:
                row_num = row_map_metric[first_key]
                if isinstance(second_data, dict):
                    for year_key, val in second_data.items():
                        if year_key in column_map_year:
                            col_letter = column_map_year[year_key]
                            cell_target = f"{col_letter}{row_num}"
                            ws[cell_target] = "" if val is None else val

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        st.success("🎉 อัปเดตข้อมูลลงตารางสำเร็จเรียบร้อยแล้ว!")
        
        st.download_button(
            label="📥ดาวน์โหลดไฟล์ Excel ที่เสร็จสมบูรณ์",
            data=excel_buffer,
            file_name="Updated_Code_CS_full_financials.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    if not (uploaded_files and excel_template):
        st.warning("⚠️ กรุณาอัปโหลดทั้งไฟล์ Template Excel และ PDF ให้ครบถ้วน")
